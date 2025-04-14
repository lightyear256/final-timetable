import os
from bs4 import BeautifulSoup
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
import re

def get_color_from_style(style):
    if not style:
        return None
    color_match = re.search(r'background-color:\s*#([0-9a-fA-F]{6})', style)
    return color_match.group(1) if color_match else None

def convert_html_to_excel(html_path, workbook, dept_name=None, term=None, section=None):
    """Convert a single HTML timetable to an Excel worksheet"""
    with open(html_path, 'r', encoding='utf-8') as f:
        soup = BeautifulSoup(f.read(), 'html.parser')
    
    # Create worksheet with appropriate name
    sheet_name = f"{dept_name}_{term}"
    if section:
        sheet_name += f"_{section}"
    sheet_name = sheet_name[:31]  # Excel sheet name length limit
    ws = workbook.create_sheet(title=sheet_name)
    
    # Set up styles
    header_fill = PatternFill(start_color='1A237E', end_color='1A237E', fill_type='solid')
    header_font = Font(color='FFFFFF', bold=True)
    break_fill = PatternFill(start_color='ECEFF1', end_color='ECEFF1', fill_type='solid')
    border = Border(left=Side(style='thin'), right=Side(style='thin'),
                   top=Side(style='thin'), bottom=Side(style='thin'))
    
    # Process timetable
    table = soup.find('table')
    if table:
        # Process headers
        headers = table.find_all('th')
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header.text.strip())
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            cell.border = border
        
        # Process rows
        for row_idx, tr in enumerate(table.find_all('tr')[1:], 2):
            cells = tr.find_all(['td'])
            col = 1
            for cell in cells:
                # Get colspan
                colspan = int(cell.get('colspan', 1))
                value = cell.text.strip()
                
                # Get cell style and color
                if 'break' in cell.get('class', []):
                    fill = break_fill
                else:
                    color_div = cell.find('div', class_='course-block')
                    if color_div and color_div.get('style'):
                        color = get_color_from_style(color_div['style'])
                        if color:
                            fill = PatternFill(start_color=color, end_color=color, fill_type='solid')
                        else:
                            fill = None
                    else:
                        fill = None
                
                # Write cell
                excel_cell = ws.cell(row=row_idx, column=col, value=value)
                if colspan > 1:
                    ws.merge_cells(start_row=row_idx, start_column=col,
                                 end_row=row_idx, end_column=col + colspan - 1)
                if fill:
                    excel_cell.fill = fill
                excel_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                excel_cell.border = border
                
                col += colspan
        
        # Add legend
        legend_table = soup.find_all('table')[1]  # Second table is the legend
        if legend_table:
            legend_start_row = len(table.find_all('tr')) + 3
            ws.cell(row=legend_start_row, column=1, value='Course Legend').font = Font(bold=True)
            
            # Process legend headers
            legend_headers = legend_table.find_all('th')
            for col, header in enumerate(legend_headers, 1):
                cell = ws.cell(row=legend_start_row + 1, column=col, value=header.text.strip())
                cell.font = Font(bold=True)
                cell.border = border
            
            # Process legend rows
            for row_idx, tr in enumerate(legend_table.find_all('tr')[1:], legend_start_row + 2):
                cells = tr.find_all('td')
                for col, cell in enumerate(cells, 1):
                    value = cell.text.strip()
                    excel_cell = ws.cell(row=row_idx, column=col, value=value)
                    if col == 2:  # Color column
                        color_div = cell.find('div', class_='legend-color')
                        if color_div and color_div.get('style'):
                            color = get_color_from_style(color_div['style'])
                            if color:
                                excel_cell.fill = PatternFill(start_color=color, end_color=color, fill_type='solid')
                    excel_cell.border = border
        
        # Adjust column widths
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = (max_length + 2)
            ws.column_dimensions[column_letter].width = min(adjusted_width, 30)
    
    return ws

def main():
    # Set up directories
    base_dir = os.path.dirname(__file__)
    html_dir = os.path.join(base_dir, 'output', 'html')
    excel_dir = os.path.join(base_dir, 'output', 'excel')
    os.makedirs(excel_dir, exist_ok=True)
    
    # Create single workbook for all timetables
    wb = Workbook()
    wb.remove(wb.active)  # Remove default sheet
    
    # Process all HTML timetables
    html_files = [f for f in os.listdir(html_dir) if f.startswith('timetable_') and f.endswith('.html')]
    
    # Sort files to group by department and semester
    html_files.sort()
    
    for html_file in html_files:
        try:
            # Extract department, semester and section info from filename
            parts = html_file.replace('.html', '').split('_')
            dept_name = parts[1]
            term = parts[3]
            section = parts[4] if len(parts) > 4 else None
            
            html_path = os.path.join(html_dir, html_file)
            convert_html_to_excel(html_path, wb, dept_name, term, section)
            print(f"Added {html_file} to combined Excel file")
        except Exception as e:
            print(f"Error converting {html_file}: {str(e)}")
    
    # Save combined Excel file
    excel_path = os.path.join(excel_dir, 'all_timetables.xlsx')
    wb.save(excel_path)
    print(f"\nAll timetables have been saved to {excel_path}")

if __name__ == "__main__":
    main()
